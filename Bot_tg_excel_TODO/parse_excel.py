import sys

import openpyxl

class ParseExcel:
    _filename = None
    _blocks = {}

    def __init__(self, filename):

        self.filename = filename

    def get_blocks(self):
        return self._blocks

    def get_time(self):
        return self.

    def write_multiple_to_excel(self, changes):
        sheet = self.workbook.active
        try:
            for row_number, column_number, value in changes:
                cell = sheet.cell(row=row_number, column=column_number)
                cell.value = value
            self.workbook.save(self.filename)
        except Exception as e:
            print(f"Ошибка при записи в Excel: {e}")

    def is_separator_row(self, row):
        """Проверяет, является ли строка разделителем блоков, проверяя только первые 12 колонок."""
        sheet = row[0].parent
        merged_ranges = []

        for cell in row[:12]:
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    merged_ranges.append(merged_range)

        if not merged_ranges:
            return False

        min_col = merged_ranges[0].min_col
        max_col = merged_ranges[0].max_col
        for merged_range in merged_ranges[1:]:
            min_col = min(min_col, merged_range.min_col)
            max_col = min(max_col, merged_range.max_col, 12)

        for col in range(min_col, max_col + 1):
            cell = row[col - 1]
            is_merged = False
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    break
            if not is_merged:
                return False

        return True

    def read_excel_and_create_dict(self):
        """Читает Excel файл, выделяет блоки и создает словарь."""
        try:
            self.workbook = openpyxl.load_workbook(self.filename)
            sheet = self.workbook.active
        except FileNotFoundError:
            print(f"Ошибка: Файл '{filename}' не найден.")
            sys.exit()
        except Exception as e:
            print(f"Произошла ошибка: {e}")
            sys.exit()

        blocks_dict = {}
        current_block = []
        current_block_name = None
        line_number = 1
        first_block_found = False

        for row in sheet.iter_rows():
            row_values = [cell.value for cell in row[:12]]

            if all(value is None for value in row_values):
                print(f"Строка {line_number} пустая, закрываем блок.")
                if current_block and first_block_found:
                    blocks_dict[current_block_name] = current_block
                    current_block = []
                    current_block_name = None
                line_number += 1
                continue

            print(f"Строка {line_number}: {row_values}")

            if self.is_separator_row(row):
                print(f"Строка {line_number} является разделителем.")
                if current_block and first_block_found:
                    blocks_dict[current_block_name] = current_block
                    current_block = []

                current_block_name = " ".join(str(x) for x in row_values if x is not None)
                first_block_found = True

            elif current_block_name is not None and first_block_found:
                print(f"Строка {line_number} не является разделителем.")
                current_block.append(row_values)

            line_number += 1

        if current_block and first_block_found:
            blocks_dict[current_block_name] = current_block
        self._blocks = blocks_dict
